#!/usr/bin/env python3
"""
build_pricing_json.py — Merge multiple vendor price books into one
pricingBook.json for the Smart MF tool.

Vendors:
  - Eagle Eye Networks (Brivo's "Video Reseller - NA1 L3" sheet)
  - Brivo Access      (Brivo's "Access Reseller - NA1 L3" sheet)
  - DoorBird          (manual extract from PDF → CSV)
  - Luxer One         (manual extract from PDF → CSV)

Schema (matches camera_markup_tool.html validatePricingBook):
  {
    schema_version: 1,
    currency: "CAD",
    updated: "<ISO date>",
    notes: "...",
    labor_rate_per_hour: <positive number>,
    items: { "<SKU>": { unit_cost: ..., msrp: ..., notes: ... } }
  }

Manual-extract CSV format (DoorBird + Luxer):
  sku,description,msrp,notes
  DB-D2101V-SSV2A,DoorBird D2101V (SSV2A finish),2369.00,Stainless V2A brushed

  unit_cost = msrp at adapter level (P0 Q3: no reseller tier in DoorBird
  or Luxer CDN books; SQ pricing rules add margin downstream).

Run:
  python3 build_pricing_json.py \\
    --brivo-ee  "source-data/April 2026 Brivo and EEN Price List NA1 Reseller L3 CDN.xlsx" \\
    --doorbird  "source-data/doorbird-extract.csv" \\
    --luxer     "source-data/luxer-extract.csv" \\
    --out       "source-data/pricingBook.json"

Missing intermediate files are SKIPPED with a warning — the converter
emits whichever vendors are present. This lets you regen the EE+Brivo
book today and add DoorBird+Luxer once the manual extracts land.
"""
import argparse
import csv
import json
import os

import openpyxl


_HERE = os.path.dirname(os.path.abspath(__file__))

# Sheet names inside the Brivo+EE combined xlsx. Two vendors, one file —
# each adapter reads its own sheet.
BRIVO_EE_SHEET_BRIVO     = 'Access Reseller - NA1 L3'
BRIVO_EE_SHEET_EAGLE_EYE = 'Video Reseller - NA1 L3'

# Effective date stamped into the merged book. Brivo+EE is the most-recent
# vendor (2026-04-01); DoorBird is 2026-01-01, Luxer is 2024-11-01.
PRICE_LIST_DATE = '2026-04-01'

# labor_rate_per_hour is REQUIRED + must be positive by the validator, but
# is NOT in any vendor price list. Placeholder; integrator edits per project.
LABOR_RATE_PLACEHOLDER = 95.00


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _parse_brivo_ee_sheet(xlsx_path, sheet_name):
    """Shared adapter for the Brivo+EE combined xlsx. Each vendor lives on
    its own sheet; column layout is identical:
      A=SKU, B=description, C=spec/sub-desc, D=msrp (CDN list),
      E=unit_cost (Reseller L3 CDN), F=notes.
    Section-header rows (text in A, empty D/E) skipped.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit('[err] sheet "%s" not in %s' % (sheet_name, xlsx_path))
    ws = wb[sheet_name]
    items = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value          # SKU
        b = ws.cell(row=r, column=2).value          # description
        c = ws.cell(row=r, column=3).value          # spec/sub-desc
        d = _num(ws.cell(row=r, column=4).value)    # list (CDN) → msrp
        e = _num(ws.cell(row=r, column=5).value)    # Reseller L3 → unit_cost
        f = ws.cell(row=r, column=6).value          # notes

        if a is None:
            continue
        sku = str(a).strip()
        if not sku:
            continue
        if e is None and d is None:
            continue

        item = {}
        if e is not None:
            item['unit_cost'] = round(float(e), 2)
        if d is not None:
            item['msrp'] = round(float(d), 2)

        note_parts = []
        if isinstance(f, str) and f.strip():
            note_parts.append(f.strip())
        if isinstance(c, str) and c.strip():
            note_parts.append(c.strip())
        desc = (str(b).strip() if isinstance(b, str) else '')
        full_note = desc
        if note_parts:
            full_note = (desc + ' — ' if desc else '') + ' · '.join(note_parts)
        if full_note:
            item['notes'] = full_note

        items[sku] = item
    return items


def parse_eagle_eye(xlsx_path):
    return _parse_brivo_ee_sheet(xlsx_path, BRIVO_EE_SHEET_EAGLE_EYE)


def parse_brivo(xlsx_path):
    return _parse_brivo_ee_sheet(xlsx_path, BRIVO_EE_SHEET_BRIVO)


def _parse_manual_csv(csv_path, vendor_label):
    """Adapter for DoorBird / Luxer / Brivo-credentials manual-extract CSVs.
    Required columns: sku, description, msrp. Optional: unit_cost, notes.
    If unit_cost is present and numeric, it OVERRIDES the msrp fallback
    (Brivo credentials CSV ships reseller pricing alongside MSRP).
    DoorBird + Luxer CSVs have no unit_cost column — row.get returns None,
    so unit_cost stays at msrp (original P0 Q3 behavior). Skips rows with
    empty SKU or non-numeric msrp; warns on the latter and on non-numeric
    unit_cost. Accepts numbers as bare, with a leading $, or with thousands
    commas (e.g. "1,234.50").
    """
    items = {}
    with open(csv_path, 'r', encoding='utf-8-sig', newline='') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            sku = (row.get('sku') or '').strip()
            if not sku:
                continue
            msrp_raw = (row.get('msrp') or '').strip()
            if not msrp_raw:
                continue
            try:
                msrp = float(msrp_raw.replace(',', '').replace('$', '').strip())
            except ValueError:
                print('[warn] %s: row "%s" has non-numeric msrp "%s" — skipped'
                      % (vendor_label, sku, msrp_raw))
                continue
            unit_cost = msrp  # Q3 default: fall back to msrp when no reseller tier known.
            uc_raw = (row.get('unit_cost') or '').strip()
            if uc_raw:
                try:
                    unit_cost = float(uc_raw.replace(',', '').replace('$', '').strip())
                except ValueError:
                    print('[warn] %s: row "%s" has non-numeric unit_cost "%s" — using msrp'
                          % (vendor_label, sku, uc_raw))
            item = {
                'unit_cost': round(unit_cost, 2),
                'msrp':      round(msrp, 2),
            }
            desc  = (row.get('description') or '').strip()
            notes = (row.get('notes') or '').strip()
            parts = []
            if desc:
                parts.append(desc)
            if notes:
                parts.append(notes)
            if parts:
                item['notes'] = ' — '.join(parts)
            items[sku] = item
    return items


def parse_doorbird(csv_path):
    return _parse_manual_csv(csv_path, 'doorbird')


def parse_luxer(csv_path):
    return _parse_manual_csv(csv_path, 'luxer')


def parse_brivo_credentials(csv_path):
    return _parse_manual_csv(csv_path, 'brivo_credentials')


def _default_path(*parts):
    return os.path.join(_HERE, *parts)


def main():
    parser = argparse.ArgumentParser(
        description='Merge vendor price books into one pricingBook.json.',
    )
    parser.add_argument(
        '--brivo-ee',
        default=_default_path('source-data',
            'April 2026 Brivo and EEN Price List NA1 Reseller L3 CDN.xlsx'),
        help='Brivo + Eagle Eye combined reseller xlsx (Brivo Access + EE Video sheets).',
    )
    parser.add_argument(
        '--doorbird',
        default=_default_path('source-data', 'doorbird-extract.csv'),
        help='DoorBird manual-extract CSV (columns: sku, description, msrp, notes).',
    )
    parser.add_argument(
        '--luxer',
        default=_default_path('source-data', 'luxer-extract.csv'),
        help='Luxer One manual-extract CSV (columns: sku, description, msrp, notes).',
    )
    parser.add_argument(
        '--brivo-credentials',
        default=_default_path('source-data', 'brivo-credentials-extract.csv'),
        help='Brivo Credentials manual-extract CSV (columns: sku, description, msrp, unit_cost, notes).',
    )
    parser.add_argument(
        '--out',
        default=_default_path('source-data', 'pricingBook.json'),
        help='Output pricingBook.json path.',
    )
    args = parser.parse_args()

    items = {}
    collisions = []
    per_vendor = {}

    # Order matters only for collision reporting. Within Brivo+EE both
    # sheets are disjoint by prefix (EN-* vs B-*); DoorBird (DB-*) and
    # Luxer (LUX-*) are vendor-namespaced. No expected overlaps.
    for vendor, fn, src in [
        ('eagle_eye',         parse_eagle_eye,         args.brivo_ee),
        ('brivo',             parse_brivo,             args.brivo_ee),
        ('doorbird',          parse_doorbird,          args.doorbird),
        ('luxer',             parse_luxer,             args.luxer),
        ('brivo_credentials', parse_brivo_credentials, args.brivo_credentials),
    ]:
        if not os.path.exists(src):
            print('[warn] skipping %s: %s not found' % (vendor, src))
            per_vendor[vendor] = 0
            continue
        try:
            vendor_items = fn(src)
        except Exception as exc:
            print('[err] %s parser failed on %s: %s' % (vendor, src, exc))
            per_vendor[vendor] = 0
            continue
        for sku, item in vendor_items.items():
            if sku in items:
                collisions.append((sku, vendor))
                print('[warn] SKU collision: %s (later vendor: %s) — overwriting'
                      % (sku, vendor))
            items[sku] = item
        per_vendor[vendor] = len(vendor_items)
        print('[ok] %s: %d items' % (vendor, len(vendor_items)))

    book = {
        'schema_version':       1,
        'currency':             'CAD',
        'updated':              PRICE_LIST_DATE,
        'notes': (
            'Merged book: Brivo (Access + Credentials) + Eagle Eye + DoorBird + Luxer One — effective '
            + PRICE_LIST_DATE + '. unit_cost = vendor reseller/dealer price where '
            'available; msrp = list. labor_rate_per_hour is a PLACEHOLDER ('
            + str(LABOR_RATE_PLACEHOLDER) + '). DoorBird + Luxer One have no '
            'reseller tier in their CDN books — unit_cost = msrp pending dealer '
            'terms; SQ section pricing rules add margin downstream.'
        ),
        'labor_rate_per_hour':  LABOR_RATE_PLACEHOLDER,
        'items':                items,
    }
    with open(args.out, 'w', encoding='utf-8') as fh:
        json.dump(book, fh, indent=2, ensure_ascii=False)
    print('[ok] wrote %s' % args.out)
    print('[ok] total: %d items across %d vendors' % (
        len(items),
        sum(1 for c in per_vendor.values() if c > 0)
    ))
    if collisions:
        print('[warn] %d SKU collision(s) — later vendor wins' % len(collisions))


if __name__ == '__main__':
    main()
